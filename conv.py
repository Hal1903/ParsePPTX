from tika import parser
import openpyxl
import os
from textblob import TextBlob

parsed = parser.from_file('order.pptx')
#print(parsed["metadata"])
#print(parsed["content"])

def check(*args):
    for file in args:
        if os.path.exists(file):
            os.remove(file)
        else:
            pass

# See next line if the current line does not have a period or a mark
#
def create(file):#*args):
    filenum=0
    parsed = file#args[filenum]
    count = 0
    f = open("test.txt", "a", encoding='utf8')
    for x in range(len(parsed["content"])):
        if parsed["content"][x] == "　":
            pass #f.write('s')
        elif parsed["content"][x] =="\n":
            f.write("\n")
        elif parsed["content"][x]==" ":
            f.write(" ")
            pass #f.write("//")
        else:
            f.write(parsed["content"][x])
            #print(x, end="")
    filenum += 1
    f.close()

def modify(file):
    f = open(file, "r", encoding="utf8")
    f2 = open("test1.txt", "a", encoding="utf8")
    removed = open("removed.txt", "a", encoding="utf8")
    count = 0
    for line in f:
        if line.isspace():
            pass
        else:
            line = line.rstrip()
            if line[-1]=="." or line[-1]=="。" or line[-1]=="?" or line[-1]=="？" or line[-1]=="!" or line[-1]=="！" or line[-1]==")" or line[-1]=="）":
                f2.write(line + "\n")
            else:
                removed.write(line+"\n")
        count += 1
    f.close()
    f2.close()
    removed.close()


def last(file):
    f = open(file, "r", encoding="utf8")
    f2 = open("test2.txt", "a", encoding="utf8")
    prec = f.readline()
    count = 0
    for line in f:
        #if line[-1]=="." or line[-1]=="。" or line[-1]=="?" or line[-1]=="？" or line[-1]=="!" or line[-1]=="！" or line[-1]==")" or line[-1]=="）":
            #f2.write(line)
        if line[0] == "(":
            prec = prec.strip()
            f2.write(prec+line)
        else:
            if (prec[0] == "(" or prec[0] == "（"): #or ():
                pass
            #elif (TextBlob(prec).detect_language()==TextBlob(line).detect_language()):
                #print(prec+" & "+line)
                #print(TextBlob(prec))
            else:
                f2.write(prec)
        prec = line
        count+=1
    f2.write(prec)
    f.close()
    f2.close()
#put prec, see current line to see if next line is needed or should be connected as a single line

def XL(pptx):
    wb = openpyxl.Workbook()
    sheet = wb.active
    r = 1
    c = 1
    count = 0
    with open(pptx, encoding="utf8") as f:
        for line in f:
            if count % 2 == 0:
                cell = sheet.cell(row=r, column=c)
                cell.value = line
                c += 1
            elif count % 2 == 1:
                cell = sheet.cell(row=r, column=c)
                cell.value = line
                r += 1
                c -= 1
            count += 1
    wb.save("order.xlsx")

def r():
    check("test.txt")
    check("test1.txt")
    check("removed.txt")
    check("test2.txt")
    check("test.txt")
    check("order.xlsx")

check("test.txt", "test1.txt", "test2.txt", "removed.txt", "test.txt", "order.xlsx")
f = "test.txt"
fx = "test1.txt"
create(parsed)
modify(f)
last(fx)
XL("test2.txt")